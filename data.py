import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# --- STEP 1: GENERATE MESSY SOURCE DATA ---
# Simulating an uncleaned, real-world raw dump file
data = {
    "Transaction_ID": [101, 102, 103, 103, 104, 105, 106, 107],  # 103 is a duplicate
    "Date": [
        "2026-07-01",
        "2026/07/02",
        "07-03-2026",
        "07-03-2026",
        "2026-07-04",
        "2026-07-05",
        "2026-07-06",
        None,
    ],
    "Region": ["North", "north", "South", "South", "East", "West", None, "West"],
    "Revenue": [2500, 1800, 3100, 3100, None, 4200, 1500, 2900],
}
df_raw = pd.DataFrame(data)
df_raw.to_csv("raw_sales_data.csv", index=False)
print("✔ Raw data file created with missing inputs, mixed casing, and duplicates.\n")


# --- STEP 2: AUTOMATED DATA CLEANING PIPELINE ---
def automated_cleaning_pipeline(file_path):
    # 1. Import raw data
    df = pd.read_csv(file_path)

    # 2. Remove exact duplicates
    df.drop_duplicates(inplace=True)

    # 3. Standardize and fix string inconsistencies (casing & text spaces)
    if "Region" in df.columns:
        df["Region"] = df["Region"].str.strip().str.capitalize()
        # Handle missing text categorical parameters
        df["Region"].fillna("Unknown", inplace=True)

    # 4. Enforce strict datetime alignment
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        # Fill missing critical dates with the logical forward fill method
        df["Date"].ffill(inplace=True)
        df["Date"] = df["Date"].dt.strftime("%Y-%m-%d")

    # 5. Handle missing numeric parameters safely
    if "Revenue" in df.columns:
        median_rev = df["Revenue"].median()
        df["Revenue"].fillna(median_rev, inplace=True)

    # Ensure absolute structural indexing
    df["Transaction_ID"] = df["Transaction_ID"].fillna(0).astype(int)

    return df


# Execute the cleaning engine
cleaned_df = automated_cleaning_pipeline("raw_sales_data.csv")
print("✔ Pipeline executed: Duplicates dropped, mixed formats unified, nulls imputed.")


# --- STEP 3: AUTOMATED REPORT GENERATION & STYLING ---
def export_styled_report(df, output_filename="Executive_Sales_Report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Performance"

    # Display sheet gridlines explicitly
    ws.views.sheetView[0].showGridLines = True

    # Build Header Title Block
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "AUTOMATED SALES SUMMARY REPORT"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Write dataframe table structure rows
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Design custom styling parameters
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    header_font = Font(name="Arial", size=11, bold=True, color="000000")
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # Style table column headers (Row 2)
    ws.row_dimensions[2].height = 25
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Format data values rows dynamically (Row 3 onwards)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            # Format numbers as standard currency, align text blocks cleanly
            if cell.column == 4:  # Revenue column formatting
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Add Dynamic KPI Total Row Calculation at footer base
    total_row_idx = ws.max_row + 1
    ws.cell(row=total_row_idx, column=3, value="Total Revenue:").font = Font(
        name="Arial", size=11, bold=True
    )
    ws.cell(row=total_row_idx, column=3).alignment = Alignment(
        horizontal="right"
    )

    revenue_formula = f"=SUM(D3:D{total_row_idx-1})"
    total_cell = ws.cell(row=total_row_idx, column=4, value=revenue_formula)
    total_cell.font = Font(name="Arial", size=11, bold=True, color="1F4E78")
    total_cell.number_format = "$#,##0.00"
    total_cell.border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )

    # Adjust dynamic tracking column width dimensions to prevent clipping layout text
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_filename)
    print(f"✔ Stylized executive spreadsheet generated: '{output_filename}'")


# Export the final presentation report
export_styled_report(cleaned_df)